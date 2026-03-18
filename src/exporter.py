"""Excel report generator and file archiver."""

import logging
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .utils import safe_filename

logger = logging.getLogger(__name__)

# Styles matching sample_report.xlsx
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="4E83FD")
DATA_FONT = Font(name="微软雅黑", size=10)
HEADER_FILL = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F8F9FB", end_color="F8F9FB", fill_type="solid")
BORDER_COLOR = "DEE0E3"
THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="medium", color="4E83FD"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

HEADERS = ["年度", "月度", "一级明细", "二级明细", "三级明细", "币种", "原币金额", "汇率", "人民币金额", "文件名称"]
COL_WIDTHS = [10, 8, 14, 22, 22, 8, 12, 10, 14, 45]


def generate_report(
    records: list[dict],
    year: str,
    month: str,
    person_name: str,
    exchange_info: dict,
    output_dir: Path,
) -> Path:
    """Generate the expense report Excel file.

    Args:
        records: Classified expense records
        year: Report year (e.g., "2026年")
        month: Report month (e.g., "3月")
        person_name: Person name
        exchange_info: Exchange rate info dict
        output_dir: Output directory

    Returns:
        Path to the generated Excel file
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "报销明细"

    # Write headers
    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = CENTER_ALIGN
        ws.column_dimensions[get_column_letter(col)].width = width

    # Write data rows
    for i, rec in enumerate(records):
        row_num = i + 2
        new_filename = build_filename(rec)

        row_data = [
            year,
            month,
            rec.get("category_l1", ""),
            rec.get("category_l2", ""),
            rec.get("category_l3", ""),
            rec.get("currency", "RMB"),
            rec.get("amount"),
            rec.get("rate_used", 1),
            rec.get("rmb_amount"),
            new_filename,
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col in (1, 2, 6):  # Center align: year, month, currency
                cell.alignment = CENTER_ALIGN
            elif col in (7, 8, 9):  # Number columns
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(val, (int, float)):
                    cell.number_format = "#,##0.00" if col != 8 else "0.0000"
            else:
                cell.alignment = LEFT_ALIGN

            # Alternate row colors
            if i % 2 == 1:
                cell.fill = ALT_ROW_FILL

    # Freeze first row
    ws.freeze_panes = "A2"

    # Add exchange rate sheet
    _add_exchange_rate_sheet(wb, exchange_info)

    # Save
    filename = f"{person_name}_{year}{month}_报销报表.xlsx"
    output_path = output_dir / safe_filename(filename)
    wb.save(output_path)
    logger.info(f"报销报表已生成: {output_path}")
    return output_path


def _add_exchange_rate_sheet(wb: openpyxl.Workbook, exchange_info: dict):
    """Add exchange rate reference sheet to the workbook."""
    ws = wb.create_sheet(title="汇率参考")

    # Title
    ws.cell(row=1, column=1, value="汇率信息").font = HEADER_FONT

    ws.cell(row=2, column=1, value="数据来源:").font = DATA_FONT
    source = exchange_info.get("source", "")
    if source == "chinamoney":
        ws.cell(row=2, column=2, value="中国人民银行外汇交易中心").font = DATA_FONT
    else:
        ws.cell(row=2, column=2, value="备用汇率（非实时）").font = DATA_FONT

    ws.cell(row=3, column=1, value="查询日期:").font = DATA_FONT
    ws.cell(row=3, column=2, value=exchange_info.get("date", "")).font = DATA_FONT

    url = exchange_info.get("url", "")
    if url:
        ws.cell(row=4, column=1, value="数据链接:").font = DATA_FONT
        ws.cell(row=4, column=2, value=url).font = DATA_FONT

    # Rate table
    rate_headers = ["货币代码", "中间价（人民币/1单位外币）", "说明"]
    for col, h in enumerate(rate_headers, 1):
        cell = ws.cell(row=6, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER

    rates = exchange_info.get("rates", {})
    for i, (code, rate) in enumerate(sorted(rates.items())):
        row_num = 7 + i
        ws.cell(row=row_num, column=1, value=code).font = DATA_FONT
        ws.cell(row=row_num, column=2, value=rate).font = DATA_FONT
        note = "每100日元" if code == "JPY" else ""
        ws.cell(row=row_num, column=3, value=note).font = DATA_FONT
        for col in range(1, 4):
            ws.cell(row=row_num, column=col).border = THIN_BORDER

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20


def build_filename(record: dict) -> str:
    """Build the standardized filename for an expense record.

    Format: 一级明细-二级明细-三级明细-币种-原币金额
    """
    if record.get("_is_credit_card"):
        # Credit card statement keeps original basename (strip subfolder)
        src = record.get("_source_file", "信用卡账单")
        return Path(src).name

    l1 = record.get("category_l1", "")
    l2 = record.get("category_l2", "")
    l3 = record.get("category_l3", "")
    currency = record.get("currency", "RMB")
    amount = record.get("amount", 0)

    # Format amount
    if amount is not None:
        try:
            amt = float(amount)
            if amt == int(amt):
                amount_str = str(int(amt))
            else:
                amount_str = str(amt)
        except (TypeError, ValueError):
            amount_str = str(amount)
    else:
        amount_str = "0"

    parts = [l1, l2, l3, currency, amount_str]
    filename = "-".join(parts)
    return safe_filename(filename)


def archive_files(
    records: list[dict],
    source_dir: Path,
    output_dir: Path,
    person_name: str,
) -> dict[str, str]:
    """Copy and rename invoice files to archive directory.

    Returns mapping of original filename → new filename.
    """
    archive_dir = output_dir / "renamed_files"
    archive_dir.mkdir(parents=True, exist_ok=True)

    # Group records by source file to handle credit card statements
    file_records: dict[str, list[dict]] = {}
    for rec in records:
        src = rec.get("_source_file", "")
        if src:
            file_records.setdefault(src, []).append(rec)

    renamed = {}
    for src_name, recs in file_records.items():
        src_path = source_dir / src_name
        if not src_path.exists():
            logger.warning(f"源文件不存在: {src_path}")
            continue

        ext = src_path.suffix

        if recs[0].get("_is_credit_card"):
            # Credit card: keep original basename (strip subfolder path)
            new_name = Path(src_name).name
        else:
            # Regular invoice: use new naming convention
            rec = recs[0]
            new_name = build_filename(rec) + ext

        dst_path = archive_dir / safe_filename(new_name)

        # Handle name conflicts
        counter = 1
        base_dst = dst_path
        while dst_path.exists():
            stem = base_dst.stem
            dst_path = base_dst.parent / f"{stem}_{counter}{ext}"
            counter += 1

        shutil.copy2(src_path, dst_path)
        renamed[src_name] = dst_path.name
        logger.info(f"归档: {src_name} → {dst_path.name}")

    return renamed
