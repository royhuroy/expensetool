"""Shared utilities for expense tool."""

import json
import re
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".heic", ".heif", ".tiff", ".tif"}
UNSAFE_CHARS = re.compile(r'[\\/:*?"<>|]')


def setup_logging(log_path: Path | None = None):
    handlers: list[logging.Handler] = [logging.StreamHandler()]
    if log_path:
        log_path.parent.mkdir(parents=True, exist_ok=True)
        handlers.append(logging.FileHandler(log_path, encoding="utf-8"))
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=handlers,
    )


def safe_filename(name: str) -> str:
    """Remove characters not allowed in Windows filenames."""
    cleaned = UNSAFE_CHARS.sub("_", name)
    return cleaned.strip().strip(".")


def load_json(path: Path) -> dict | list:
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.loads(f.read().replace("\r\n", "\n"))


def save_json(path: Path, data: dict | list):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def scan_invoice_files(invoices_dir: Path) -> list[Path]:
    """Scan directory recursively for supported invoice/receipt files."""
    files = []
    if not invoices_dir.exists():
        return files
    for f in sorted(invoices_dir.rglob("*")):
        if f.is_file() and not f.name.startswith(".") and f.suffix.lower() in SUPPORTED_EXTENSIONS:
            files.append(f)
    return files


def list_subfolders(base_dir: Path) -> list[Path]:
    """List all subfolders (recursively) under base_dir, including base_dir itself."""
    folders = [base_dir]
    if base_dir.exists():
        for d in sorted(base_dir.rglob("*")):
            if d.is_dir() and not d.name.startswith("."):
                folders.append(d)
    return folders


def is_credit_card_statement(filename: str) -> bool:
    return "信用卡账单" in filename or "credit card" in filename.lower()


def credit_card_mode(filename: str) -> str:
    """Determine credit card processing mode from filename.
    Returns 'all' or 'highlight'.
    """
    name_lower = filename.lower()
    if "所有明细" in filename or "识别所有" in filename or "all" in name_lower:
        return "all"
    if "highlight" in name_lower or "高亮" in filename:
        return "highlight"
    return "all"  # default to all


def extract_categories_from_sample(sample_path: Path) -> list[dict]:
    """Extract unique category combinations from sample Excel."""
    import openpyxl
    if not sample_path.exists():
        return []
    wb = openpyxl.load_workbook(sample_path, read_only=True)
    ws = wb.active
    categories = []
    seen = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        l1, l2, l3 = row[2], row[3], row[4]
        if l1:
            l2_str = str(l2) if l2 else ""
            l3_str = str(l3) if l3 else ""
            key = (str(l1), l2_str, l3_str)
            if key not in seen:
                seen.add(key)
                categories.append({"l1": str(l1), "l2": l2_str, "l3": l3_str})
    wb.close()
    return categories


def extract_sample_styles(sample_path: Path) -> dict:
    """Extract Excel formatting info from sample report."""
    import openpyxl
    if not sample_path.exists():
        return {}
    wb = openpyxl.load_workbook(sample_path)
    ws = wb.active
    styles = {
        "headers": [],
        "col_widths": {},
        "font_name": "微软雅黑",
        "font_size": 10,
        "header_color": "4E83FD",
        "header_bold": True,
    }
    for cell in ws[1]:
        if cell.value:
            styles["headers"].append(cell.value)
            col_letter = cell.column_letter
            if ws.column_dimensions[col_letter].width:
                styles["col_widths"][col_letter] = ws.column_dimensions[col_letter].width
            if cell.font:
                styles["font_name"] = cell.font.name or "微软雅黑"
                styles["font_size"] = cell.font.size or 10
    wb.close()
    return styles
